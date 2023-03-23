import openpyxl
from pathlib import Path
from sys import argv

SCENARIOROW = 9
PROGTCENTERROW = 11
DEPTROW = 12

YEARCOL = 4
PERIODCOL = 6

#outputPath = str(Path.cwd()) + 'output.xlsx'
output = openpyxl.Workbook()
outputSheet = output.create_sheet(title='Formatted')

output.active = outputSheet

outputRow = 1

def outputValues(keyString, progtCenter, department, year, scenario, period, budget, inputSheet):
    # writing parsed values to the new output sheet
    global outputSheet, outputRow

    outputSheet.cell(row = outputRow, column = 1).value = keyString
    outputSheet.cell(row = outputRow, column = 2).value = progtCenter
    outputSheet.cell(row = outputRow, column = 3).value = department
    outputSheet.cell(row = outputRow, column = 4).value = year
    outputSheet.cell(row = outputRow, column = 5).value = scenario
    outputSheet.cell(row = outputRow, column = 6).value = period
    outputSheet.cell(row = outputRow, column = 7).value = budget
    
    output.save(filename='Output.xlsx')
    outputRow += 1

# search leftmost column for keywords/identifiers and append to output workbook
# in order: keyString, Progt Center, dept, year, scenario, period, amount
def selectRowInformation(keyString, rowNum, activeSheet):
        print("Keyword hit: select, print, and append row to output")
        
        # search the columns to the right until getting an amount

        # TODO rethink colNum approach, not really intuitive at all
        colNum = 7
        for col in activeSheet.iter_cols(min_row = rowNum, min_col = 8, max_col = activeSheet.max_column, values_only = True):
            colNum += 1
            for budget in col:
                if budget is not None:
                    # number hit
                    # now can retrieve column specific information
                    progtCenter = activeSheet.cell(PROGTCENTERROW,colNum).value
                    department = activeSheet.cell(DEPTROW,colNum).value
                    scenario = activeSheet.cell(SCENARIOROW,colNum).value
                    year = activeSheet.cell(rowNum, YEARCOL).value
                    period = activeSheet.cell(rowNum, PERIODCOL).value
                    print(keyString, progtCenter, department, year, scenario, period, budget)
                    # write these values to output workbook
                    outputValues(keyString, progtCenter, department, year, scenario, period, budget, activeSheet)
def main():
    # load input workbook that is to be parsed
    input = openpyxl.load_workbook(str(Path.cwd()) + '\\' + argv[1])

    # for each worksheet in input
    for tab in input.sheetnames:
        print(tab)
        tab = input.active
        # perform parsing operations
        row = 0
        for value in tab.iter_rows(0, tab.max_row, 0, 1, values_only = True):
            row += 1
            if value[0] is not None:
                # case statement for constant time lookup opposed to if else
                # should probably move to its own module, makes debugging this file complicated
                # is there a way to make this faster? see how match is carried out in Python

                match value[0]:
                    case "401104 - Bottle Deposits":
                        selectRowInformation(value[0], row, tab)
                    case "401107 - Pharmaceutical Service Income":
                        selectRowInformation(value[0], row, tab)
                    case "401108 - Gift Card Commission":
                        selectRowInformation(value[0], row, tab)
                    case "401109 - Lottery Income":
                        selectRowInformation(value[0], row, tab)
                    case "401110 - Money Order Fee":
                        selectRowInformation(value[0], row, tab)
                    case "401111 - Financial Services Income":
                        selectRowInformation(value[0], row, tab)
                    case "401112 - Transportation and Entertainment Income":
                        selectRowInformation(value[0], row, tab)
                    case "401113 - Rental Equipment Income":
                        selectRowInformation(value[0], row, tab)
                    case "401114 - Government Commissions Income":
                        selectRowInformation(value[0], row, tab)
                    case "401115 - Other Retail Sales - Agent":
                        selectRowInformation(value[0], row, tab)
                    case "401183 - Sales Discounts - Pre Perpetual Inventor":
                        selectRowInformation(value[0], row, tab)
                    case "401184 - Merch Promo- Pre Perpetual Inventory":
                        selectRowInformation(value[0], row, tab)
                    case "401185 - Bottle Deposits - Pre Perpetual inventor":
                        selectRowInformation(value[0], row, tab)
                    case "614101 - Supplies":
                        selectRowInformation(value[0], row, tab)
                    case "614102 - Front End Bags":
                        selectRowInformation(value[0], row, tab)
                    case "614103 - Packaging":
                        selectRowInformation(value[0], row, tab)
                    case "614104 - Unit Price Tags":
                        selectRowInformation(value[0], row, tab)
                    case "614106 - Cleaning Products":
                        selectRowInformation(value[0], row, tab)
                    case "614107 - Company Clothing":
                        selectRowInformation(value[0], row, tab)
                    case "614108 - Print Shop Charges":
                        selectRowInformation(value[0], row, tab)
                    case "614109 - Pallets":
                        selectRowInformation(value[0], row, tab)
                    case "614131 - Cash Differences":
                        selectRowInformation(value[0], row, tab)
                    case "614132 - Money Handling and Financial Fees":
                        selectRowInformation(value[0], row, tab)
                    case "614134 - Bank Charges":
                        selectRowInformation(value[0], row, tab)
                    case "614135 - Credit/Debit/EBT Card Fees":
                        selectRowInformation(value[0], row, tab)
                    case "614136 - Petty Cash-Paid Out":
                        selectRowInformation(value[0], row, tab)
                    case "614137 - Anti Fraud Service fees":
                        selectRowInformation(value[0], row, tab)
                    case "614138 - Credit Card Charge backs":
                        selectRowInformation(value[0], row, tab)
                    case "614139 - Vendor Coupon Differences":
                        selectRowInformation(value[0], row, tab)
                    case "614141 - Charge Sales Transaction fees":
                        selectRowInformation(value[0], row, tab)
                    case "614161 - Repair & Maintenance - Parts":
                        selectRowInformation(value[0], row, tab)
                    case "614171 - Equipment - Non-Capital":
                        selectRowInformation(value[0], row, tab)
                    case "614181 - Allocation Other Operating Expense":
                        selectRowInformation(value[0], row, tab)
                    case "614301 - Software Licenses":
                        selectRowInformation(value[0], row, tab)
                    case "614302 - Hardware Purchases":
                        selectRowInformation(value[0], row, tab)
                    case "614303 - Data Communications":
                        selectRowInformation(value[0], row, tab)
                    case "614306 - Telephone":
                        selectRowInformation(value[0], row, tab)
                    case "614502 - Postage":
                        selectRowInformation(value[0], row, tab)
                    case "614503 - Office Supplies":
                        selectRowInformation(value[0], row, tab)
                    case "614504 - Meeting and Events Expense":
                        selectRowInformation(value[0], row, tab)
                    case "614511 - Market Research":
                        selectRowInformation(value[0], row, tab)
                    case "614512 - Communications":
                        selectRowInformation(value[0], row, tab)
                    case "614513 - Customer Satisfaction expenses":
                        selectRowInformation(value[0], row, tab)
                    case "614521 - Charitable Donations":
                        selectRowInformation(value[0], row, tab)
                    case "614543 - Opex Savings":
                        selectRowInformation(value[0], row, tab)
                    case "614544 - Settlement Expenses":
                        selectRowInformation(value[0], row, tab)
                    case "614545 - Quality Assurance Testing":
                        selectRowInformation(value[0], row, tab)
                    case "614546 - Closing Cost Adjustment":
                        selectRowInformation(value[0], row, tab)
                    case "614602 - Royalties":
                        selectRowInformation(value[0], row, tab)
                    case "614603 - Penalties":
                        selectRowInformation(value[0], row, tab)
                    case "614604 - Exchange Rate Gain & Loss":
                        selectRowInformation(value[0], row, tab)
                    case "614611 - Bad Debts":
                        selectRowInformation(value[0], row, tab)
                    case "614621 - Loss Expenses":
                        selectRowInformation(value[0], row, tab)
                    case "614701 - Costs of Fixed Assets to be Capitalized":
                        selectRowInformation(value[0], row, tab)
                    case "614804 - Training Costs":
                        selectRowInformation(value[0], row, tab)
                    case "614901 - Other Operational Expenses":
                        selectRowInformation(value[0], row, tab)
                    case "614905 - Other Operational Expenses- Non Recurrin":
                        selectRowInformation(value[0], row, tab)
                    case "614906 - Other Operational Expenses-Restructuring":
                        selectRowInformation(value[0], row, tab)
                    case "621103 - Overtime Wages":
                        selectRowInformation(value[0], row, tab)
                    case "621104 - Wages Special Days Off":
                        selectRowInformation(value[0], row, tab)
                    case "621105 - Training/Meeting Wages and Salaries":
                        selectRowInformation(value[0], row, tab)
                    case "621106 - Lump Sum Wages":
                        selectRowInformation(value[0], row, tab)
                    case "621107 - Other Direct Wages and Salaries":
                        selectRowInformation(value[0], row, tab)
                    case "621108 - Other Indirect Wages and Salaries":
                        selectRowInformation(value[0], row, tab)
                    case "621111 - Base Wages and Salaries - Exempt (Salary":
                        selectRowInformation(value[0], row, tab)
                    case "621112 - Base Wages and Salaries - Non Exempt (Ho":
                        selectRowInformation(value[0], row, tab)
                    case "621113 - Project Wages":
                        selectRowInformation(value[0], row, tab)
                    case "621114 - Pre Opening Salaries":
                        selectRowInformation(value[0], row, tab)
                    case "621115 - Vacation Coverage Salaries":
                        selectRowInformation(value[0], row, tab)
                    case "621202 - Vacation Pay":
                        selectRowInformation(value[0], row, tab)
                    case "621203 - Holiday Pay":
                        selectRowInformation(value[0], row, tab)
                    case "621301 - Sick Pay":
                        selectRowInformation(value[0], row, tab)
                    case "621401 - Employee Allowances":
                        selectRowInformation(value[0], row, tab)
                    case "621701 - Short Term Incentive Bonus":
                        selectRowInformation(value[0], row, tab)
                    case "621702 - Other Bonuses":
                        selectRowInformation(value[0], row, tab)
                    case "621704 - Severance Payment":
                        selectRowInformation(value[0], row, tab)
                    case "621951 - Allocation of Wages and Salaries":
                        selectRowInformation(value[0], row, tab)
                    case "642101 - Electricity":
                        selectRowInformation(value[0], row, tab)
                    case "642102 - Water/Sewer":
                        selectRowInformation(value[0], row, tab)
                    case "642103 - Utilities - Other":
                        selectRowInformation(value[0], row, tab)
                    case "642104 - Common Area Maintenance Utilities":
                        selectRowInformation(value[0], row, tab)
                    case "642105 - Gas":
                        selectRowInformation(value[0], row, tab)
                    case "642106 - Allocation - Utilities":
                        selectRowInformation(value[0], row, tab)
                    case "643101 - Recycling Income":
                        selectRowInformation(value[0], row, tab)
                    case "643102 - Pallet Income":
                        selectRowInformation(value[0], row, tab)
                    case "643103 - Bottle Handling Income":
                        selectRowInformation(value[0], row, tab)
                    case "643104 - Unloading Income":
                        selectRowInformation(value[0], row, tab)
                    case "643105 - Backhaul Income":
                        selectRowInformation(value[0], row, tab)
                    case "643106 - Other Income":
                        selectRowInformation(value[0], row, tab)
                    case "643107 - Other Income - Tax":
                        selectRowInformation(value[0], row, tab)
                    case "643201 - Coupon Redemption":
                        selectRowInformation(value[0], row, tab)
                    case "643211 - Fixed Asset Gain Other":
                        selectRowInformation(value[0], row, tab)
                    case "643212 - Coinstar Revenue Share Income":
                        selectRowInformation(value[0], row, tab)
                    case "643213 - Returned Check Service Charges Income":
                        selectRowInformation(value[0], row, tab)
                    case "643214 - Energy Efficiency Income":
                        selectRowInformation(value[0], row, tab)
                    case "643215 - Independent Fees Income":
                        selectRowInformation(value[0], row, tab)
                    case "643216 - Settlement Miscellaneous income":
                        selectRowInformation(value[0], row, tab)
                    case "643217 - In Store Tenant Income":
                        selectRowInformation(value[0], row, tab)
                    case "643301 - Vendor Compliance Violation Income":
                        selectRowInformation(value[0], row, tab)
                    case "643401 - Allocation Other Income":
                        selectRowInformation(value[0], row, tab)
                    case "643402 - Segment Insurance Allocation Income":
                        selectRowInformation(value[0], row, tab)
                    case "675410 - Rounding Differences":
                        selectRowInformation(value[0], row, tab)
                    case "699999 - WBS Conversion":
                        selectRowInformation(value[0], row, tab)
                    case "PL200102 - Other Retail sales":
                        selectRowInformation(value[0], row, tab)
                    case "PL21210 - Online sales":
                        selectRowInformation(value[0], row, tab)
                    case "PL24799 - Net sales":
                        selectRowInformation(value[0], row, tab)
                    case "PL400101 - Cost of Product":
                        selectRowInformation(value[0], row, tab)
                    case "PL400102 - Cost of Product Other":
                        selectRowInformation(value[0], row, tab)
                    case "PL40080 - Write-downs of inventories":
                        selectRowInformation(value[0], row, tab)
                    case "PL400801 - Shrink":
                        selectRowInformation(value[0], row, tab)
                    case "PL400802 - Obsolete Stock":
                        selectRowInformation(value[0], row, tab)
                    case "PL400803 - Other write downs of inventory":
                        selectRowInformation(value[0], row, tab)
                    case "PL400804 - Reclamations":
                        selectRowInformation(value[0], row, tab)
                    case "PL40110 - Vendor allowances":
                        selectRowInformation(value[0], row, tab)
                    case "PL40299 - Gross profit":
                        selectRowInformation(value[0], row, tab)
                    case "PL41100 - Wages and salaries":
                        selectRowInformation(value[0], row, tab)
                    case "PL411001 - Wages and salaries including benefits":
                        selectRowInformation(value[0], row, tab)
                    case "PL411002 - Vacation/Holiday/Sick Leave":
                        selectRowInformation(value[0], row, tab)
                    case "PL411003 - Bonus/Training/Other Indirect pay":
                        selectRowInformation(value[0], row, tab)
                    case "PL4110031 - Bonus":
                        selectRowInformation(value[0], row, tab)
                    case "PL4110032 - Other Indirect Pay":
                        selectRowInformation(value[0], row, tab)
                    case "PL41210 - Pension expenses":
                        selectRowInformation(value[0], row, tab)
                    case "PL41300 - Health & medical expenses":
                        selectRowInformation(value[0], row, tab)
                    case "PL41400 - Share-based compensation expenses":
                        selectRowInformation(value[0], row, tab)
                    case "PL41500 - Other employee costs":
                        selectRowInformation(value[0], row, tab)
                    case "PL41700 - Contracted personnel":
                        selectRowInformation(value[0], row, tab)
                    case "PL41800 - Capitalization of labor costs":
                        selectRowInformation(value[0], row, tab)
                    case "PL41899 - Labor Costs":
                        selectRowInformation(value[0], row, tab)
                    case "PL42259 - Advertising":
                        selectRowInformation(value[0], row, tab)
                    case "PL43100 - Hired services":
                        selectRowInformation(value[0], row, tab)
                    case "PL431001 - Repairs and Maintenance":
                        selectRowInformation(value[0], row, tab)
                    case "PL4310011 - Building/Lighting/Plumbing (through Service C":
                        selectRowInformation(value[0], row, tab)
                    case "PL4310012 - Refrigeration (through Service Channel)":
                        selectRowInformation(value[0], row, tab)
                    case "PL4310013 - Other Repairs and Maintenace (through Service":
                        selectRowInformation(value[0], row, tab)
                    case "PL4310014 - Repairs and Maintenance (Non-Service Channel)":
                        selectRowInformation(value[0], row, tab)
                    case "PL431002 - Common Area Maintenance":
                        selectRowInformation(value[0], row, tab)
                    case "PL431003 - IT Maintenance (Store Automation)":
                        selectRowInformation(value[0], row, tab)
                    case "PL431004 - Other Services":
                        selectRowInformation(value[0], row, tab)
                    case "PL431005 - Transaction Fees":
                        selectRowInformation(value[0], row, tab)
                    case "PL431006 - Cleaning Services":
                        selectRowInformation(value[0], row, tab)
                    case "PL431007 - Security":
                        selectRowInformation(value[0], row, tab)
                    case "PL43200 - Utilities":
                        selectRowInformation(value[0], row, tab)
                    case "PL43300 - Other income":
                        selectRowInformation(value[0], row, tab)
                    case "PL43400 - Other operational expenses":
                        selectRowInformation(value[0], row, tab)
                    case "PL434001 - Supplies":
                        selectRowInformation(value[0], row, tab)
                    case "PL434002 - Other expenses":
                        selectRowInformation(value[0], row, tab)
                    case "PL4340021 - Travel":
                        selectRowInformation(value[0], row, tab)
                    case "PL4340022 - Transportation":
                        selectRowInformation(value[0], row, tab)
                    case "PL4340023 - IT Services":
                        selectRowInformation(value[0], row, tab)
                    case "PL4340024 - Money handling":
                        selectRowInformation(value[0], row, tab)
                    case "PL4340025 - Other expenses":
                        selectRowInformation(value[0], row, tab)
                    case "PL434003 - Taxes and Licenses":
                        selectRowInformation(value[0], row, tab)
                    case "PL434004 - Insurance":
                        selectRowInformation(value[0], row, tab)
                    case "PL43800 - Capitalization of other costs":
                        selectRowInformation(value[0], row, tab)
                    case "PL43900 - Allocation income / expense":
                        selectRowInformation(value[0], row, tab)
                    case "PL43950 - Intercompany cross charges":
                        selectRowInformation(value[0], row, tab)
                    case "PL44899 - Net Rent":
                        selectRowInformation(value[0], row, tab)
                    case "PL45600 - Depreciation and amortization":
                        selectRowInformation(value[0], row, tab)
                    case "PL45899 - Subtotal cost center expenses":
                        selectRowInformation(value[0], row, tab)
                    case _:
                        continue

if __name__ == '__main__':
    main()